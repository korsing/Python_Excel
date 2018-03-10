# 엑셀을 다루기 위해 필요한 파이썬 모듈
import xlsxwriter
import openpyxl
import os
   

def prepare_excel():
    # 최종 출석부 엑셀파일 오픈
    excel_01 = openpyxl.load_workbook("출석부01.xlsx")
    sheet_01 = excel_01["Sheet1"]
    # LC 분반 기재 엑셀파일 오픈
    excel_LC = openpyxl.load_workbook("LC별배정내역.xlsx")
    sheet_LC = excel_LC["GEDT01401"]

    # 종합 출석부 생성
    excel_Final = xlsxwriter.Workbook("FINAL_01.xlsx")
    sheet_Final = excel_Final.add_worksheet("종합")
    
    writeExcel(sheet_Final, 0, 0, "순번")
    writeExcel(sheet_Final, 0, 1, "계열 / 학과")
    writeExcel(sheet_Final, 0, 2, "LC 분반")
    writeExcel(sheet_Final, 0, 3, "학번")
    writeExcel(sheet_Final, 0, 4, "이름")

    student_list_LC = []
    counter1 = 0
    # LC파일에서 학번만 모아서 리스트로 저장
    for rownum_LC in range(1, 150):
        student_id_LC = sheet_LC.cell(row=rownum_LC, column=5).value
        student_LC = sheet_LC.cell(row=rownum_LC, column=4).value
        if(type(student_id_LC) == int):
            student_list_LC.append(student_id_LC)
            counter1 += 1
    
    # 최종 출석 파일에서 학번을 가져오기
    counter = 0
    for rownum_01 in range(1, 230):
        student_id_01 = sheet_01.cell(row=rownum_01, column=13).value
        if(type(student_id_01) == int):
            if(student_id_01 in student_list_LC):
                counter += 1
                dept = sheet_01.cell(row=rownum_01, column=8).value
                LC = student_LC
                studentid = student_id_01
                name = sheet_01.cell(row=rownum_01, column=15).value
                print(counter, dept, LC, studentid, name)
    
    excel_01.close()
    excel_Final.close()

def writeExcel(file, row, column, content):
    file.write(row, column, content)

    
# 메인 함수
os.chdir("C:\\Users\\Joshua Y. S. Jung\\Downloads\\컴퓨팅사고")
prepare_excel()

