# 엑셀을 다루기 위해 필요한 파이썬 모듈
import xlsxwriter
import openpyxl
import os

# 플래닝
# 출석부 파일에서 학번을 따와서 LC 분반이 기재되어있는 엑셀 파일에서 검색
# 맞다면 LC 분반 엑셀에서 LC를 따와서 출석부 파일에 복사 붙여넣기.
   

def prepare_excel():
    # 교수님이 주신 02분반 출석부 오픈
    final_list = openpyxl.load_workbook("출석부02.xlsx")
    # 그 중에서 첫 번째 시트 선택
    loadsheet = final_list["Sheet1"]

    LCfile = openpyxl.load_workbook("LC별배정내역.xlsx")
    LCsheet = LCfile["GEDT01402"]
    # 저장할 최종 파일 이름 설정 후
    destination = xlsxwriter.Workbook("FINAL_02.xlsx")
    # 시트 설정해서 엑셀파일 생
    savesheet = destination.add_worksheet("종합")
    # 최종 엑셀 파일에 구분 셀 입력
    writeExcel(savesheet, 0, 0, "순번")
    writeExcel(savesheet, 0, 1, "계열 / 학과")
    writeExcel(savesheet, 0, 2, "LC 분반")
    writeExcel(savesheet, 0, 3, "학번")
    writeExcel(savesheet, 0, 4, "이름")


    total = []
    counter = 0
    for rownum1 in range(10, 230):
        student_id = loadsheet.cell(row=rownum1, column=3).value
        if(type(student_id) == int):
            total.append(student_id)
            counter += 1
    
    for rownum2 in range(2, 149):
        id = LCsheet.cell(row=rownum2, column=5).value
        if(type(id) == int):
            
    
    final_list.close()
    destination.close()

def writeExcel(file, row, column, content):
    file.write(row, column, content)

    
# 메인 함수
os.chdir("C:\\Users\\Joshua Y. S. Jung\\Downloads\\컴퓨팅사고")
prepare_excel()

